# input: ChartRecordRepository、BirthChartService 与排盘请求模型。
# output: 档案记录的业务编排，以及列表/详情响应映射与结果页兼容数据，并统一空值为“未知”展示。
# pos: 后端档案服务层。
# 一旦我被更新务必更新我的开头注释以及所属文件夹的 md
from __future__ import annotations

import json
import tempfile
import uuid
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment

from ..repositories import ChartRecordFilters, ChartRecordRepository
from ..schemas.requests import BirthChartRequest
from ..schemas.responses import (
    ApiCaseChartsViewModel,
    ApiCaseViewModel,
    BannerViewModel,
    BirthChartApiResponse,
    ChartRecordDetailResponse,
    ChartRecordListItem,
    ChartRecordListResponse,
)
from .chart_service import BirthChartService, ChartRequestError

UNKNOWN_DISPLAY = "未知"


class ChartRecordService:
    def __init__(
        self,
        repository: ChartRecordRepository,
        chart_service: BirthChartService,
    ) -> None:
        self._repository = repository
        self._chart_service = chart_service

    def create_record(self, user_id: int, request: BirthChartRequest) -> ChartRecordDetailResponse:
        record_data, case_data_list = self._build_persistence_payload(request)
        record_id = self._repository.create_record(user_id, record_data, case_data_list)
        record = self._repository.get_record(user_id, record_id)
        if record is None:
            raise ChartRequestError("排盘记录保存成功后无法读取")
        return self._map_record_detail(record)

    def update_record(self, user_id: int, record_id: int, request: BirthChartRequest) -> ChartRecordDetailResponse:
        record_data, case_data_list = self._build_persistence_payload(request, preserve_created_at=True)
        updated = self._repository.update_record(user_id, record_id, record_data, case_data_list)
        if not updated:
            raise ChartRequestError(f"排盘记录不存在: {record_id}")
        record = self._repository.get_record(user_id, record_id)
        if record is None:
            raise ChartRequestError(f"排盘记录不存在: {record_id}")
        return self._map_record_detail(record)

    def delete_record(self, user_id: int, record_id: int) -> None:
        deleted = self._repository.delete_record(user_id, record_id)
        if not deleted:
            raise ChartRequestError(f"排盘记录不存在: {record_id}")

    def list_records(
        self,
        user_id: int,
        page: int,
        page_size: int,
        *,
        name: str | None = None,
        birth_date: str | None = None,
        region_id: str | None = None,
        zi_hour_type: str | None = None,
        has_lunar_leap_case: bool | None = None,
        digit_string: str | None = None,
        chart_type: str | None = None,
    ) -> ChartRecordListResponse:
        filters = ChartRecordFilters(
            name=name,
            birth_date=birth_date,
            region_id=region_id,
            zi_hour_type=zi_hour_type,
            has_lunar_leap_case=has_lunar_leap_case,
            digit_string=digit_string,
            chart_type=chart_type,
        )
        total, rows = self._repository.list_records(user_id, filters, page, page_size)
        return ChartRecordListResponse(
            items=[self._map_record_list_item(row) for row in rows],
            total=total,
            page=page,
            pageSize=page_size,
        )

    def get_record(self, user_id: int, record_id: int) -> ChartRecordDetailResponse:
        record = self._repository.get_record(user_id, record_id)
        if record is None:
            raise ChartRequestError(f"排盘记录不存在: {record_id}")
        return self._map_record_detail(record)

    def export_records(self, user_id: int, *, name: str | None = None, digit_string: str | None = None) -> tuple[Path, str]:
        filters = ChartRecordFilters(name=name, digit_string=digit_string)
        rows = self._repository.list_records_for_export(user_id, filters)
        if not rows:
            raise ChartRequestError("当前没有可导出的档案")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "档案导出"
        worksheet.append(["档案名", "性别", "阳历生日", "阳格", "阳格缺漏", "农历生日", "阴格", "阴格缺漏", "时辰"])

        wrap_alignment = Alignment(vertical="top", wrap_text=True)
        single_line_alignment = Alignment(vertical="top")

        for row in rows:
            detail = self._repository.get_record(user_id, row["id"])
            if detail is None or not detail["cases"]:
                continue

            export_values = self._build_export_row(detail)
            worksheet.append(export_values)
            current_row = worksheet.max_row
            for column_index in range(1, 10):
                worksheet.cell(row=current_row, column=column_index).alignment = (
                    wrap_alignment if column_index in {4, 5, 6, 7, 8} else single_line_alignment
                )

        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].width = 10
        worksheet.column_dimensions["C"].width = 16
        worksheet.column_dimensions["D"].width = 18
        worksheet.column_dimensions["E"].width = 18
        worksheet.column_dimensions["F"].width = 18
        worksheet.column_dimensions["G"].width = 18
        worksheet.column_dimensions["H"].width = 18
        worksheet.column_dimensions["I"].width = 10

        temp_dir = Path(tempfile.gettempdir()) / "nine_grid_record_exports"
        temp_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"档案批量导出_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        file_path = temp_dir / f"{uuid.uuid4().hex}.xlsx"
        workbook.save(file_path)
        return file_path, file_name

    @staticmethod
    def _build_grid_data(chart_type: str, chart: object) -> dict[str, object]:
        chart_dict = chart.model_dump()
        return {
            "chart_type": chart_type,
            "digit_string": chart_dict["digitString"],
            "missing_digits": chart_dict["missingDigits"],
            "missing_attributes": chart_dict["missingAttributes"],
            "missing_count": chart_dict["missingCount"],
            "po": chart_dict["po"],
            "po_raw": chart_dict["poRaw"],
            "main_soul": chart_dict["mainSoul"],
            "sub_soul": chart_dict["subSoul"],
            "half_supplement": "" if chart_dict["halfSupplement"] == UNKNOWN_DISPLAY else chart_dict["halfSupplement"],
            "grid_cells_json": json.dumps(chart_dict["cells"], ensure_ascii=False),
            "grid_snapshot_json": json.dumps(chart_dict, ensure_ascii=False),
        }

    def _build_persistence_payload(
        self,
        request: BirthChartRequest,
        preserve_created_at: bool = False,
    ) -> tuple[dict[str, object], list[dict[str, object]]]:
        computed_chart = self._chart_service.build_computed_birth_chart(request)
        timestamp = datetime.now(UTC).replace(microsecond=0).isoformat(sep=" ")
        response = computed_chart.response
        raw_cases = computed_chart.raw_result["cases"]
        region = computed_chart.region

        record_data = {
            "name": response.summary.name,
            "gender": "" if request.gender == UNKNOWN_DISPLAY else request.gender,
            "birth_date": request.birthDate,
            "birth_time": request.birthTime,
            "input_datetime_text": " ".join(filter(None, [request.birthDate, request.birthTime])),
            "region_id": request.regionId,
            "province_name": region.province_name if request.regionId else "",
            "city_name": region.city_name if request.regionId else "",
            "district_name": region.district_name if request.regionId else "",
            "longitude": region.longitude if request.regionId else 0,
            "zi_hour_type": self._map_zi_hour_type(response.summary.ziHourType),
            "case_count": len(response.cases),
            "true_solar_datetime": self._storage_or_empty(response.summary.trueSolarDatetimeText),
            "true_solar_shichen": self._storage_or_empty(response.summary.trueSolarShichen),
            "has_lunar_leap_case": any(case.metrics.lunarIsLeapMonth for case in response.cases),
            "source_payload_json": json.dumps(request.model_dump(), ensure_ascii=False),
            "created_at": timestamp,
            "updated_at": timestamp,
            "preserve_created_at": preserve_created_at,
        }

        case_data_list: list[dict[str, object]] = []
        for case, raw_case in zip(response.cases, raw_cases, strict=True):
            case_data_list.append(
                {
                    "case_index": case.index,
                    "case_label": case.label,
                    "date_relation": raw_case["date_relation"],
                    "solar_birthday": case.metrics.solarBirthday,
                    "lunar_birthday": case.metrics.lunarBirthday,
                    "lunar_birthday_display": case.metrics.lunarBirthdayDisplay,
                    "age": case.metrics.age,
                    "true_solar_shichen": self._storage_or_empty(case.metrics.trueSolarShichen),
                    "lunar_is_leap_month": case.metrics.lunarIsLeapMonth,
                    "grids": [
                        self._build_grid_data("yang", case.charts.yang),
                        self._build_grid_data("yin", case.charts.yin),
                    ],
                }
            )

        return record_data, case_data_list

    @staticmethod
    def _map_zi_hour_type(value: str) -> str:
        return {
            "非子时": "none",
            "前子时": "front",
            "后子时": "back",
        }[value]

    @staticmethod
    def _map_zi_hour_type_display(value: str) -> str:
        return {
            "none": "非子时",
            "front": "前子时",
            "back": "后子时",
        }[value]

    @staticmethod
    def _build_banners_from_cases(cases: list[ApiCaseViewModel], zi_hour_type: str) -> list[BannerViewModel]:
        banners: list[BannerViewModel] = []
        if zi_hour_type != "非子时":
            banners.append(
                BannerViewModel(
                    type="warning",
                    code="zi-hour",
                    title="子时双方案提示",
                    description="当前结果存在两套有效方案，请结合方案切换一起查看，不要只读取第一套。",
                )
            )
        if any(case.metrics.lunarIsLeapMonth for case in cases):
            banners.append(
                BannerViewModel(
                    type="info",
                    code="lunar-leap",
                    title="农历闰月提示",
                    description="当前结果存在农历闰月方案，闰月补 1 会参与阴格正式计算，并可能影响最终串与灵魂结构。",
                )
            )
        return banners

    @classmethod
    def _build_export_row(cls, record: dict[str, object]) -> list[str]:
        cases = record["cases"]
        yang_grids: list[str] = []
        yang_missing_values: list[str] = []
        lunar_birthdays: list[str] = []
        yin_grids: list[str] = []
        yin_missing_values: list[str] = []

        for case_row in cases:
            grid_map = {grid["chart_type"]: cls._parse_snapshot(grid["grid_snapshot_json"]) for grid in case_row["grids"]}
            yang_grids.append(grid_map["yang"]["digitString"])
            yang_missing_values.append(grid_map["yang"]["missingDigits"])
            lunar_birthdays.append(case_row["lunar_birthday"])
            yin_grids.append(grid_map["yin"]["digitString"])
            yin_missing_values.append(grid_map["yin"]["missingDigits"])

        return [
            cls._display_name(record.get("name")),
            cls._display_or_unknown(record.get("gender")),
            str(record.get("birth_date") or ""),
            "\n".join(yang_grids),
            "\n".join(yang_missing_values),
            "\n".join(lunar_birthdays),
            "\n".join(yin_grids),
            "\n".join(yin_missing_values),
            cls._format_shichen_for_export(record.get("true_solar_shichen")),
        ]

    @staticmethod
    def _storage_or_empty(value: str) -> str:
        return "" if value == UNKNOWN_DISPLAY else value

    @staticmethod
    def _display_or_unknown(value: object) -> str:
        normalized = str(value or "").strip()
        return normalized or UNKNOWN_DISPLAY

    @staticmethod
    def _display_name(value: object) -> str:
        normalized = str(value or "").strip()
        return normalized or "未命名档案"

    @staticmethod
    def _format_shichen_for_export(value: object) -> str:
        normalized = str(value or "").strip()
        if not normalized:
            return UNKNOWN_DISPLAY
        if normalized == UNKNOWN_DISPLAY or normalized.endswith("时"):
            return normalized
        return f"{normalized}时"

    @staticmethod
    def _parse_snapshot(value: str | None) -> dict[str, object]:
        return json.loads(value or "{}")

    @classmethod
    def _build_region_text_from_row(cls, row: dict[str, object]) -> str:
        region_text = " ".join(
            part
            for part in [
                str(row.get("province_name") or "").strip(),
                str(row.get("city_name") or "").strip(),
                str(row.get("district_name") or "").strip(),
            ]
            if part
        )
        return region_text or UNKNOWN_DISPLAY

    def _map_record_list_item(self, row: dict[str, object]) -> ChartRecordListItem:
        region_text = self._build_region_text_from_row(row)
        detail = self._repository.get_record(row["user_id"], row["id"])
        if detail is None or not detail["cases"]:
            raise ChartRequestError(f"排盘记录不存在: {row['id']}")
        first_case = detail["cases"][0]
        grid_map = {grid["chart_type"]: self._repository.parse_json_text(grid["grid_snapshot_json"]) for grid in first_case["grids"]}
        return ChartRecordListItem(
            id=row["id"],
            name=row["name"],
            gender=self._display_or_unknown(row["gender"]),
            birthDate=row["birth_date"],
            birthTime=row["birth_time"],
            regionId=row["region_id"],
            regionText=region_text,
            ziHourType=self._map_zi_hour_type_display(row["zi_hour_type"]),
            caseCount=row["case_count"],
            hasLunarLeapCase=bool(row["has_lunar_leap_case"]),
            trueSolarDatetimeText=self._display_or_unknown(row["true_solar_datetime"]),
            trueSolarShichen=self._display_or_unknown(row["true_solar_shichen"]),
            createdAt=row["created_at"],
            firstCaseSolarBirthday=first_case["solar_birthday"],
            firstCaseLunarBirthday=first_case["lunar_birthday"],
            firstCaseYangDigitString=grid_map["yang"]["digitString"],
            firstCaseYangMissingDigits=grid_map["yang"]["missingDigits"],
            firstCaseYinDigitString=grid_map["yin"]["digitString"],
            firstCaseYinMissingDigits=grid_map["yin"]["missingDigits"],
        )

    def _map_record_detail(self, row: dict[str, object]) -> ChartRecordDetailResponse:
        cases: list[ApiCaseViewModel] = []
        for case_row in row["cases"]:
            chart_map = {}
            for grid_row in case_row["grids"]:
                snapshot = self._repository.parse_json_text(grid_row["grid_snapshot_json"])
                chart_map[grid_row["chart_type"]] = snapshot
            charts = ApiCaseChartsViewModel(
                yang=chart_map["yang"],
                yin=chart_map["yin"],
            )
            cases.append(
                ApiCaseViewModel(
                    index=case_row["case_index"],
                    label=case_row["case_label"],
                    metrics={
                        "solarBirthday": case_row["solar_birthday"],
                        "lunarBirthday": case_row["lunar_birthday"],
                        "lunarBirthdayDisplay": case_row["lunar_birthday_display"],
                        "age": case_row["age"],
                        "trueSolarShichen": self._display_or_unknown(case_row["true_solar_shichen"]),
                        "lunarIsLeapMonth": bool(case_row["lunar_is_leap_month"]),
                    },
                    charts=charts,
                )
            )

        zi_hour_type_display = self._map_zi_hour_type_display(row["zi_hour_type"])
        region_text = self._build_region_text_from_row(row)
        banners = self._build_banners_from_cases(cases, zi_hour_type_display)
        return ChartRecordDetailResponse(
            id=row["id"],
            name=row["name"],
            gender=self._display_or_unknown(row["gender"]),
            birthDate=row["birth_date"],
            birthTime=row["birth_time"],
            inputDatetimeText=row["input_datetime_text"],
            regionId=row["region_id"],
            regionText=region_text,
            longitude=row["longitude"],
            ziHourType=zi_hour_type_display,
            caseCount=row["case_count"],
            hasLunarLeapCase=bool(row["has_lunar_leap_case"]),
            trueSolarDatetimeText=self._display_or_unknown(row["true_solar_datetime"]),
            trueSolarShichen=self._display_or_unknown(row["true_solar_shichen"]),
            createdAt=row["created_at"],
            updatedAt=row["updated_at"],
            banners=banners,
            cases=cases,
        )
